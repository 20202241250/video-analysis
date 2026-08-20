# -*- coding: utf-8 -*-
import os
import json
import cv2
import numpy as np
import shutil
import base64
import random
import re
import logging
from functools import wraps
from flask import Flask, render_template, request, jsonify, redirect, url_for, session

# 假设 auth 模块中已实现 require_login
from auth import auth_login, require_login

# 如果需要 Excel 读取，确保安装 openpyxl
from openpyxl import load_workbook

# 初始化日志（可根据需要配置日志级别与格式）
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

app = Flask(__name__)
app.secret_key = 'wefwef43543t435y54y345wy45wy564y'  # 用于会话加密

# OpenAI 客户端配置（根据实际情况调整）
from openai import OpenAI

vl_client = OpenAI(
    api_key="EMPTY",
    base_url="http://10.16.1.7:8088/v1",
    timeout=360.0
)

client = OpenAI(
    api_key=os.getenv('DASHSCOPE_API_KEY'),
    base_url="https://dashscope.aliyuncs.com/compatible-mode/v1",
)

# 配置上传和存储路径
app.config['UPLOAD_FOLDER'] = './uploads'
app.config['KEYFRAMES_FOLDER'] = './static/keyframes'
app.config['ALLOWED_EXTENSIONS'] = {'mp4'}

# 确保目录存在
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['KEYFRAMES_FOLDER'], exist_ok=True)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        # 简单的用户名和密码验证
        if username == 'dut' and password == 'dut123456':
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='Invalid credentials')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route('/', methods=['GET'])
@require_login
def index():
    return render_template('index.html')

@app.route('/api/upload', methods=['POST'])
@require_login
def handle_upload():
    """视频上传接口"""
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
    if not (file and allowed_file(file.filename)):
        return jsonify({"error": "File type not allowed"}), 409
    try:
        filename = file.filename
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        # 为每次上传创建唯一的关键帧目录
        import uuid
        session_id = str(uuid.uuid4())
        output_folder = os.path.join(app.config['KEYFRAMES_FOLDER'], session_id)
        os.makedirs(output_folder, exist_ok=True)

        # 提取关键帧（使用 decord 及向量化处理）
        extract_key_frames(filename, file_path, output_folder, start_time=0, end_time=None, frames_per_second=10, similarity_threshold=300, desired_keyframes=40)

        # 获取关键帧URL列表
        image_urls = get_keyframe_urls(output_folder)

        # 检查关键帧数量
        if len(image_urls) < 4 or len(image_urls) > 2000:
            return jsonify({"error": "The number of keyframes must be between 4 and 2000."}), 400

        # 清理上传的视频文件
        os.remove(file_path)

        data = {'keyframes': image_urls}

        return jsonify({'code': 200, 'data': data})
    except Exception as e:
        logging.exception("Error in handle_upload")
        return jsonify({"error": str(e)}), 500

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def get_keyframe_urls(keyframes_dir):
    """获取关键帧的URL列表"""
    image_urls = []
    relative_path = os.path.relpath(keyframes_dir, 'static')
    for file in os.listdir(keyframes_dir):
        if file.lower().endswith(('.jpg', '.jpeg')):
            url = f'/static/{relative_path}/{file}'
            image_urls.append(url)
    # 使用正则从文件名中提取序号进行排序
    def extract_index(fname):
        m = re.search(r'_keyframe_(\d+)\.jpg$', fname)
        return int(m.group(1)) if m else 0
    image_urls = sorted(image_urls, key=lambda url: extract_index(os.path.basename(url)))
    return image_urls

def extract_key_frames(filename, video_path, output_folder, start_time=0, end_time=None,
                       frames_per_second=50, similarity_threshold=300, desired_keyframes=30):
    """
    使用 decord 批量读取视频帧，计算灰度差异，根据差异挑选关键帧，
    并调整最终关键帧数量为 desired_keyframes（默认50）。
    """
    from decord import VideoReader, cpu
    vr = VideoReader(video_path, ctx=cpu(0))
    fps = vr.get_avg_fps()
    total_frames = len(vr)

    start_frame = int(start_time * fps)
    end_frame = total_frames if end_time is None else min(int(end_time * fps), total_frames)

    # 计算采样间隔，获取所有采样帧索引
    interval = max(1, int(fps / frames_per_second))
    indices = list(range(start_frame, end_frame, interval))
    if not indices:
        indices = [start_frame]

    # 批量读取帧，decord 返回 RGB 格式
    frames = vr.get_batch(indices).asnumpy()

    # 转换为灰度图（利用加权求和）
    gray_frames = np.dot(frames[..., :3], [0.2989, 0.5870, 0.1140])

    # 计算相邻帧灰度差的平均值
    diffs = np.abs(np.diff(gray_frames, axis=0))
    avg_diffs = np.mean(diffs, axis=(1, 2))

    # 根据阈值选择关键帧，初始帧始终保留
    key_indices = [0]
    for i, diff in enumerate(avg_diffs, start=1):
        if diff > similarity_threshold:
            key_indices.append(i)

    # 使用两步控制策略保证最终关键帧数为 desired_keyframes
    total_sampled = len(frames)
    if len(key_indices) < desired_keyframes:
        # 如果筛选结果太少，则从未选中的帧中补充
        remaining = list(set(range(total_sampled)) - set(key_indices))
        remaining.sort()
        needed = desired_keyframes - len(key_indices)
        if len(remaining) >= needed:
            extra = sorted(random.sample(remaining, needed))
        else:
            extra = remaining  # 如果剩余帧不足，则全部补上
        key_indices = sorted(key_indices + extra)
    elif len(key_indices) > desired_keyframes:
        # 如果筛选结果太多，则随机采样 desired_keyframes 个
        key_indices = sorted(random.sample(key_indices, desired_keyframes))
    # 如果刚好等于 desired_keyframes，则直接使用

    # 保存关键帧图片
    key_frame_files = []
    for count, idx in enumerate(key_indices):
        frame = frames[idx]
        # 转换为 BGR 格式后保存
        bgr_frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
        key_frame_filename = f"{output_folder}/{filename}_keyframe_{count}.jpg"
        cv2.imwrite(key_frame_filename, bgr_frame)
        key_frame_files.append(key_frame_filename)

    print(f"Extracted {len(key_frame_files)} key frames.")
    return key_frame_files


def get_image_base64(image_path):
    """将图片转换为base64格式"""
    with open(image_path, 'rb') as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

def excel_prompt(file_path, sheet_name='师幼空间指标'):
    """
    读取 Excel 表格数据并转换为字符串（此处简单返回所有单元格内容，实际应用中可根据需求调整）
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    content = []
    for row in ws.iter_rows(values_only=True):
        content.append("\t".join([str(cell) for cell in row if cell is not None]))
    return "\n".join(content)

def call_behavior_analysis(video_description, model, user_prompt):
    """调用百炼模型服务-行为分析模型"""
    # 获取 Excel 数据文本（建议提前缓存该数据）
    ws_text = excel_prompt("/home/dut/video-analysis-test/师幼空间幼儿活动关键经验指标.xlsx")
    prompt = (
        # f"{user_prompt}\n"
        # "# 输入数据\n"
        # "以下是视频的详细描述：\n"
        # f"{video_description}\n\n"
        # "- 《分析评价》要求:\n"
        # "  …（此处为具体要求，保持不变）\n\n"
        # f"{ws_text}\n"
        # "# 输出格式示例\n"
        # "……\n"
        # "# 强制验证\n"
        # "……\n"
        # "# 输出格式\n"
        # "直接输出修改好的幼儿观察记录，不要输出其他信息。\n"
        f"{user_prompt}\n"
        "# 输入数据\n"
        "以下是视频的详细描述：\n"
        f"{video_description}\n\n"
        "- 《分析评价》要求:\n"
        "- 分析评价的依据性:指对幼儿行为的分析评价都是依据于本次的观察记录。避免脱离观察记录本身“未记先评”、 “先入为主”、“无中生有”。\n"
        "- 分析评价的多元性:指分析评价会关注幼儿发展的多个领域、多个方面，全面的、多元的分析幼儿的行为表现。结合幼儿行为从多个领域、多个维度的关键发展指标进行条分缕析地分析。\n"
        "- 分析评价的深入性:对观察结果的分析评价不停留在泛泛的“领域”，能够深入到具体的领域发展目标、典型行为表现上。将幼儿的行为表现聚焦于幼儿某个领域的某个方面的某个发展目标上，进行具体、深入的分析。\n"
        "- 分析评价的发展性:对观察结果的分析是发展性的，能够比较准确地描绘出幼儿下一步发展的方向和目标。\n"
        "- 《支持策略》要求\n"
        "- 支持策略的具体性:后续教育策略都是围绕此次观察活动的具体的、可操作的教育建议。避免宽泛的、笼统的教育策略。要结合分析结果以及本次活动或观察对象的实际情况制定具体的支持策略。\n"
        "- 支持策略的发展性:提出的后续支持策略具有较强的发展性，能够准确基于幼儿下一步发展的方向和目标（最近发展区）提出教育建议。\n"
        "- 支持策略的关联性:观察记录中对幼儿行为的描述可以很好地对应对幼儿行为的分析，对幼儿行为的分析也可以很好地对应支持策略。\n"
        "根据《支持策略》要求在发展评估后进行支持策略分点分析，不要显示具体性等词语\n"
        
        "根据视频内容自动匹配6个最相关分析维度,对幼儿进行详细的评价和分析\n"
        # "分析维度必须严格使用以下一级指标名称，不要修改任何字符\n"
        # "   - 兴趣状态，角色游戏，建构游戏，美工创造，益智游戏，图书阅读，数学认知，科学探究，自然探究，表演游戏，沙水游戏，回顾分享，操节运动，动作技能，运动素质，自理能力，生活习惯，社会交往，学习品质\n"
        "# 能力分析输出格式示例\n"
        "- 社会交往：\n"
        "- 小雅:小雅在攀爬成功后主动表达自己的喜悦，吸引了小杰的关注，两人形成了短暂的互动。这种行为表明小雅具备一定的社交主动性，能够通过语言和动作引发他人关注。\n"
        "- 小辉:小辉踢球的行为吸引了小雅和小杰的参与，三人迅速融入共同游戏，表现出良好的合作倾向和团队意识。他们能够快速适应群体活动，说明其社会交往能力较强。\n"
        "- 小杰：虽然未直接参与游戏，但持续观察其他孩子的活动，反映出他对同伴行为的兴趣，然而缺乏主动性可能表明他在社交互动方面需要进一步引导\n"
        "- 兴趣状态：\n"
        "- 小雅：对攀爬和踢球表现出浓厚的兴趣，始终处于积极的情绪状态，未因失败或困难而退缩，说明她的游戏兴趣较高且情绪调控能力较好。\n"
        "- 小辉：在踢球过程中保持专注，情绪稳定，能够持续投入游戏，体现了良好的情绪调控能力。\n\n"
        
        "# 发展评估生成规范\n"
        "1. 总体评价四要素：\n"
        "   - 情境嵌入：'在户外足球游戏中，当...时表现出...'\n"
        "   - 能力关联：'这体现了其在XX领域XX能力的发展'\n"
        "   - 群体参照：'相较于同龄幼儿的典型表现...'\n"
        "   - 成长叙事：'从最初...到如今...显示出...的进步轨迹'\n\n"
        "2. 发展建议三维模型：\n"
        "   - 环境适配：'在建构区增加斜面轨道材料，促进...'\n"
        "   - 师幼互动：'采用'挑战阶梯'策略，逐步增加...'\n"
        "   - 家园协同：'建议家庭开展周末探索日，记录...'\n\n"
        "3. 语言优化要求：\n"
        "   - 使用教育叙事句式：'我们注意到...' '值得关注的是...'\n"
        "   - 避免评估术语堆砌：将'达到水平四'转化为'能够持续15分钟专注搭建'\n"
        "   - 保留专业性的自然表达：'这种双侧协调性发展，为书写准备奠定基础'\n\n"
        "# 发展评估输出示例\n"
        "总体评价:\n"
        "   - 在持续30分钟的户外建构游戏中，穿红色上衣的男孩展现出令人欣喜的进步。当积木塔意外倒塌时（情境嵌入），他迅速调整呼吸并说'我们再搭个更稳的吧'（情绪调控-水平四），这种挫折应对能力已超过多数同龄幼儿（群体参照）。相较上月容易放弃的表现（成长叙事），穿红色上衣的男孩在坚持性和问题解决方面取得显著发展（能力关联）。\n\n"
        "发展建议\n"
        "1. 环境升级：\n"
        "   - 在建构区投放地震模拟板，拓展抗倒塌探究（环境适配）\n"
        "   - 增设工程日志本，记录建构思路（师幼互动）\n"
        "2. 家庭延伸：\n"
        "   - 利用家庭坐垫设计多米诺挑战（家园协同）\n"
        "   - 拍摄建构过程视频进行回顾反思（家园协同）\n"
        "3. 个性化支持：\n"
        "   - 邀请担任'安全监理员'，强化责任意识（师幼互动）\n"
        "   - 每周设置'超级挑战时间'，逐步延长专注时长（环境适配）\n"
        
        "根据视频内容自动匹配《师幼空间幼儿活动关键经验指标》5个最相关二级指标,对幼儿进行等级评分\n"
        "# 限制\n"
        "1. 输入数据若未检测到幼儿（如空场景/纯物品画面），立即返回错误：[ERROR] 未检测到幼儿活动\n"
        "2. 评分需严格基于视频中幼儿的具体表现，参考以下标准：\n"
        f"{ws_text}\n"
        "3. 评分要求：\n"
        "   - 根据视频内容自动匹配表格中5个最相关二级指标（需满足：①出现频次≥3次 ②涉及≥2个幼儿），采用1-5分制\n"
        "   - 示例匹配逻辑：建构游戏→建构技能；角色游戏→角色认知\n"
        "   - 需体现领域间差异，禁止出现全5分或全3分\n"
        "4. 对不同幼儿有尽可能详细的描述,针对不同幼儿进行发展评估和建议,评分依据需覆盖60%以上幼儿行为\n"
        "5. 评分标准必须严格使用以下二级指标名称，不要修改任何字符\n"
        "   - 游戏兴趣，情绪调控\n"
        "   - 角色认知，角色分配，角色表现，主题情节，材料认知与使用\n"
        "   - 建构主题，建构技能，材料认知，材料使用，记录表征，拼插技能\n"
        "   - 感受欣赏，美的表现力，材料认知，材料使用，问题解决，工具使用\n"
        "   - 观察能力，理解能力，解决问题，精细动作\n"
        "   - 倾听习惯，讲述能力，描述能力，阅读习惯，符号意识，前书写能力\n"
        "   - 数的认知，数的运算，量的感知，数量关系，几何形体，空间方位，空间视觉化，模式，分类，时间认知\n"
        "   - 观察能力，猜想假设，实验操作，信息采集与记录，分析思考，结果的表达交流\n"
        "   - 观察能力，种植与饲养，测量与记录，分析思考，交流与表达，认识动植物，爱护动植物\n"
        "   - 作品理解，主题情节，角色分配，角色认知和行为，材料选择与使用\n"
        "   - 探究能力，塑形能力，问题解决，材料选择与使用\n"
        "   - 材料收整，记录与表征，回顾能力，叙事能力，描述能力，倾听习惯\n"
        "   - 动作准确性，动作合拍性，动作动力性\n"
        "   - 走，跑，立定跳远，单脚跳，双脚连续跳，钻，爬，滚动与滚翻，脚踢球，拍篮球，骑行\n"
        "   - 平衡性，协调性，灵敏性，力量，耐力\n"
        "   - 穿脱衣物，清洁整理，饮食自理，时间管理\n"
        "   - 问候习惯，个人卫生，作息规律，安全意识，环保意识\n"
        "   - 社交兴趣，融入游戏，友好相处，遵守规则，合作倾向，处理矛盾，尊重他人，自我认同，诚实守信，履行职责，热爱集体\n"
        "   - 主动性，目的性，计划性，专注性，独立性，坚持性，反思性，创造性\n\n"
        # "## 精准匹配规则",
        # "指标名称强制校验：",
        # "   - 必须严格使用表格二级指标名称",
        "# 输出格式示例\n"
        "- 友好相处：4分\n"
        "- 穿红色上衣的男孩主动将足球递给穿黄色外套的男孩，表现出分享和合作的行为。穿黄色外套和绿色外套的男孩共同参与踢球游戏，表现出良好的合作能力和团队意识，能与同伴友好相处，愿意与人分享\n"
        "- 情绪调控：5分\n"
        "- 大部分幼儿能够保持积极情绪状态，并在活动中展现良好适应性。\n\n"
        
        "# 强制验证\n"
        "执行以下检查（任一不通过立即报错）：\n"
        "1. 统计出现'幼儿'/'孩子'等关键词次数 ≥3次\n"
        "2. 至少描述2个不同幼儿的行为\n"
        "3. 包含具体动作描述（如搭建/奔跑/绘画）\n\n"
        "# 错误处理\n"
        "未通过验证时返回：[ERROR] 未检测到有效幼儿活动\n"
        "# 输出格式\n"
        "直接输出修改好的幼儿观察记录，不要输出其他信息。\n"
    )

    completion = client.chat.completions.create(
        model=model,  # 使用用户选择的模型
        messages=[{'role': 'user', 'content': prompt}],
    )
    return completion.model_dump_json()

@app.route('/api/vision-analysis', methods=['POST'])
@require_login
#单线程
def handle_vision_analysis():
    """视觉分析接口"""
    try:
        data = request.json
        keyframes = data.get('keyframes')
        prompt = data.get('prompt', '')
        model = data.get('model', 'deepseek-r1')  # 默认使用 deepseek-r1

        # 将URL转换为本地文件路径，并转换为 base64 格式
        base64_images = []
        total_frames = len(keyframes)
        # 使用正则安全排序：提取文件名中的关键帧序号
        def extract_index(url):
            fname = os.path.basename(url)
            m = re.search(r'_keyframe_(\d+)\.jpg$', fname)
            return int(m.group(1)) if m else 0
        keyframes = sorted(keyframes, key=extract_index)

        for i, url in enumerate(keyframes, 1):
            try:
                parts = url.split('/static/')
                if len(parts) == 2:
                    relative_path = parts[1]
                    local_path = os.path.normpath(os.path.join('static', relative_path))
                    if os.path.exists(local_path):
                        logging.info(f"处理图片 {i}/{total_frames}: {local_path}")
                        image_base64 = get_image_base64(local_path)
                        base64_images.append(f"data:image/jpeg;base64,{image_base64}")
                    else:
                        logging.warning(f"文件不存在: {local_path}")
            except Exception as e:
                logging.exception(f"处理图片出错 {url}: {str(e)}")
                continue

        if not base64_images:
            return jsonify({'code': 400, 'data': '无法处理图片文件'}), 400

        # 按批次调用视觉模型接口，每批50张图片
        batch_size = 20
        results = []
        # results = ""
        for i in range(0, len(base64_images), batch_size):
            response_raw = call_multimodal_conversation(base64_images[i:i + batch_size])
            response = json.loads(response_raw)
            if 'error' in response:
                return jsonify({'code': 400, 'data': response['error']['message']})
            results.append(response['choices'][0]['message']['content'])

        video_analysis_text = "".join(results)
        data_raw = call_text_completions(video_analysis_text)
        final_description = json.loads(data_raw)['choices'][0]['message']['content']

        return jsonify({
            'code': 200,
            'data': {
                'description': final_description  # 返回视频描述用于后续分析
            }
        })

    except Exception as e:
        logging.exception("Error in handle_vision_analysis")
        return jsonify({'error': str(e)}), 500



@app.route('/api/behavior-analysis', methods=['POST'])
@require_login
def handle_behavior_analysis():
    """行为分析接口"""
    try:
        data = request.json
        description = data.get('description')
        prompt = data.get('prompt', '')
        model = data.get('model', 'deepseek-r1')

        # 调用行为分析模型
        behavior_analysis_json = call_behavior_analysis(description, model, prompt)
        behavior_analysis = json.loads(behavior_analysis_json)['choices'][0]['message']['content']

        return jsonify({
            'code': 200,
            'data': {
                'description': description,
                'analysis': behavior_analysis
            }
        })
    except Exception as e:
        logging.exception("Error in handle_behavior_analysis")
        return jsonify({'error': str(e)}), 500

def call_multimodal_conversation(keyframes):
    """调用百炼模型服务-视觉模型"""
    prompt = (
        "# 角色\n"
        "你是一位全球顶尖的幼儿行为观察学者，具备细致观察力和敏锐的逻辑分析能力，擅长客观记录和描述每个画面中的细微差别。\n\n"
        "# 核心指令\n"
        "严格基于输入的关键帧图像序列进行视觉分析，输出幼儿活动观察报告。若检测到非幼儿园场景或无3-6岁幼儿立即终止并报错。\n\n"
        "# 任务描述\n"
        "给你一个视频片段的多张关键帧图片和文字描述（各观察对象的特征，语言描写或者其他相关的信息），请按照要求输出《幼儿观察记录》，除此之外不需要任何解释和回复。\n"
        "- 输出每张图片的画面信息，包括人物、物体、动作、文字、字幕、一句话总结等。\n"
        "- 把每张图片的信息串联起来，使用时间连接词（随后/同时/接着）衔接不同帧的关联动作，生成视频的详细概述，还原该片段的剧情。\n\n"
        "# 限制\n"
        "- 分析范围严格限定于提供的视频子片段，不涉及视频之外的任何推测或背景信息。\n"
        "- 总结时需严格依据视频内容，不可添加个人臆测或创意性内容。不要出现教师的意识以及心理活动，过程实录最后不要出现总结性或者评价性的语句\n"
        "- 视频可能拍摄于幼儿园的某个区域，包括：建构区 、美工区 、益智区、 表演区、 科学区、 图书区、 娃娃家、 自然角、数学区、沙水区、户外运动区等，请给出最可能的地点\n"
        "- 保持对所有视频元素（尤其是文字和字幕）的高保真还原，避免信息遗漏或误解。\n\n"
        "# 输出格式\n"
        "直接按照任务目标里即可，先输出每张图片的描述，再串联起来输出整个视频片段的剧情，不少于500个中文汉字文本长度。\n"
        "按时间顺序呈现每张图像中：\n"
        "- 主要角色（描述特征）及其正在进行的具体动作\n"
        "- 关键物体状态（位置/形态/使用方式）\n"
        "- 环境文字/标识内容（精确到字符）\n\n"
        "用3-5个自然段描述：\n"
        "1. 起始场景：时间/地点/初始人物配置\n"
        "2. 过程演进：关键交互动作与物体状态变化\n"
        "3. 阶段成果：当前片段结束时呈现的状态\n\n"
        "# 强制检测规则\n"
        "若画面中未出现3-6岁幼儿或幼儿园立即报错\n"
        "# 错误格式\n"
        "发现上述情况时直接返回：[ERROR] 未检测到幼儿活动\n\n"
        # " # 角色"
        # " 你是一位全球顶尖的幼儿行为观察学者，拥有极其细致的观察力和敏锐的逻辑分析能力，专注于客观记录画面中的所有细微差别，包括人物、物体、动作、文字和字幕。你不得加入任何主观情感、评价、总结或推测。"
        # " "
        # " # 任务描述"
        # " 你将获得一个幼儿视频片段的多张关键帧图片以及相关文字描述（涵盖观察对象特征、语言描写、字幕等信息）。请严格依据以下要求生成《幼儿观察记录》，输出内容中禁止出现除记录之外的任何解释或回复。"
        # " "
        # " 1. **逐张图片描述**"
        # "    - 对每张图片，客观记录画面中出现的所有元素：必须包含人物（仅限3-6岁幼儿）、物体、动作、文字、字幕等。"
        # "    - 用一句话对图片整体内容做准确概括。"
        # "    - 输出你对该图片观察的思维链，描述你如何逐步识别、解析和归纳画面信息的过程，但不得包含教师内心情感、评价或总结性语句。"
        # " "
        # " 2. **视频片段剧情概述**"
        # "    - 将所有图片的描述和思维链依次串联，形成一个完整、详细的视频片段剧情概述，力求全面还原画面中呈现的情景与发展过程。"
        # "    - 剧情概述文字总量不得少于500个中文汉字，语言必须准确、客观，仅基于画面信息，不包含任何推测或创意性内容。"
        # " "
        # " # 限制条件"
        # " - **分析范围限定**：仅限于提供的视频子片段内容，禁止出现视频之外的背景信息或推测。"
        # " - **客观记录**：所有描述必须严格基于画面中的文字、字幕及视觉信息，禁止出现教师主观意识、心理活动、评价或建议。"
        # " - **地点判断**：视频拍摄地点可能为幼儿园内各区域（如建构区、美工区、益智区、表演区、科学区、图书区、娃娃家、自然角、数学区、沙水区、户外运动区等），请根据画面环境元素给出最可能的场景。"
        # " - **高保真还原**：对所有视频元素（尤其是文字和字幕）需高保真还原，杜绝信息遗漏或误解。"
        # " - **强制检测规则**：若画面中未出现3-6岁幼儿或幼儿园环境，立即返回错误信息：[ERROR] 未检测到幼儿活动。"
        # " "
        # " # 输出格式要求"
        # " - **顺序与结构**"
        # "   -先按图片顺序逐张输出每张图片的描述及思维链，格式须以“图片X：”开头；"
        # "   -再依次串联各图片描述，输出一个完整的视频片段剧情概述，文字总量不得少于500个中文汉字。"
        # " - **禁止内容**：禁止出现任何与任务无关的解释、总结、教师视角或建议性文字；"
        # " - **语言风格**：文字表达必须客观、详实、忠实于视频片段信息，禁止主观色彩或评价性用语。"
        # " "
        # " 请严格按照以上提示生成《幼儿观察记录》，确保满足所有要求。"

    )

    messages = [{
        "role": "user",
        "content": [
            {"type": "text", "text": prompt}
        ]
    }]

    for keyframe in keyframes:
        messages[0]['content'].append({"type": "image_url", "image_url": {"url": keyframe}})


    response = vl_client.chat.completions.create(
        model="Qwen2-5-VL-32B-Instruct",
        # model="Qwen2-5-VL-72B-Instruct-AWQ",
        messages=messages,
        temperature=0.01,
        max_tokens=8024,
    )
    return response.model_dump_json()

def call_text_completions(video_analysis_text):
    """调用百炼模型服务-文本模型"""
    prompt = (
        # "# 角色\n"
        # "你是一位专业的幼儿园园长，擅长修改幼师的幼儿观察记录。\n\n"
        # "# 核心指令\n"
        # "基于视频关键帧分析生成符合《幼儿园观察记录》标准的客观叙述，若检测到非3-6岁幼儿活动立即终止并报错。\n\n"
        # "# 输入数据\n"
        # "以下是视频关键帧分析结果：\n"
        # f"{video_analysis_text}\n\n"
        # "# 输出格式\n"
        # "请直接输出修改好的幼儿观察记录，不要输出其他信息。\n"
        "# 角色\n"
        "你是一个专业的幼儿园园长，擅长修改幼师的幼儿观察记录。\n\n"
        "# 核心指令\n"
        "基于视频关键帧分析生成符合《幼儿园观察记录》标准的客观叙述，若检测到非3-6岁幼儿活动立即终止并报错。\n\n"
        "# 任务目标\n"
        "请你结合输入数据串联、还原出整个幼儿活动的详细情况。\n\n"
        "# 限制\n"
        "1. 如出现语法上错误，或逻辑不通，请直接修改\n"
        "2. 在描述中，如果包含台词，可能会出现说话者与其所说内容不匹配的情况。因此，必须根据剧情的进展，准确判断每段台词的真实说话者\n"
        "3. 如果记录中无台词，请根据视频音频文字为其匹配台词\n"
        "4. 修改后的故事请适当保留对人物、动作、区域的描写\n"
        "5. 文字由多次模型调用生成，每段描述图片序号都从1开始，请注意合并使其更具逻辑性，但要避免主观描写\n"
        "6. 结合人物外观特点，如果有外观相近的人物是同一个角色。因此，需要将不同描述中的人物角色统一，不要用孩子、幼儿等词语统一代指。\n"
        "7. 活动实录中只要白描描写，不要出现评价性或者总结性的内容，不要出现模糊词语（好像、似乎、可能、仿佛）\n\n"
        "# 输入数据\n"
        "## 资料一：幼儿观察记录\n"
        f"{video_analysis_text}\n\n"
        "- 《过程实录》要求\n"
        "- 观察目的与观察内容紧密相连，即便没有明确的观察目的，也能从观察内容中看到教师观察的重点，避免“流水账”现象的出现。要注重：客观性、完整性、细致性三个重点来描写\n"
        "- 记录的是事实，而不是观点。详细记录幼儿的语言、神态、动作、表情等。客观描述幼儿的行为表现，很少使用主观词汇或对幼儿的想法或意愿进行主观猜测。\n"
        "- 完整记录幼儿行为发生的前因、过程和结果。前因：幼儿行为发生的前因和背景（环境、人等） 是什么？过程: 幼儿的行为表现，包括幼儿说了什么、做了什么？结果: 幼儿行为的结果，包括幼儿的感受和收获、 周围人的反应是什么？\n"
        "- 对观察重点作细致、生动的描写，非重点的内容简略描述 。避免大概、笼统地描述幼儿的关键行为表现。围绕观察重点进行拆解，对相关的关键行为进行详细的观察和记录。不要出现以下词语：“好像、可能、似乎、仿佛”\n"
        "- 避免语言表达机械化，模仿人类（幼儿园老师）的写作风格，具备描述自然性与表达丰富度\n\n"
        "# 输出格式\n"
        "根据《过程实录》要求直接输出修改好的幼儿观察记录，不要输出其他信息，不要出现评价性或者总结性的内容。\n"
        "# 输出格式示例\n"
        "- 观察对象：穿着白色带有花卉图案上衣的幼儿（暂称小明）\n"
        "- 过程实录：\n"
        "- 区域游戏开始了，小明来到美工区开始制作“棒棒糖”，她拿着一盒白色的超轻粘土，从里面取出一小块黏土，单手揉捏了一下随后又用两只小手合并揉了一下，便开始往一根木根的顶端粘贴，粘上去后她来回按压，使黏土和木根的粘合更牢固一些。粘贴好之后她将“白色棒棒糖”并排放在另一跟已经做好的“棒棒糖”旁边，准备继续制作下一个。\n"
        "- 她拿起黏土准备再取出一些，忽然发现没有木棍了，于是她起身从材料架上又拿来一根木棍。她继续左手端着黏土盒，右手多次从里面挖出黏土，并单手里来回揉捏使黏土变圆润，她一边取黏土一边跟老师说：“她做的是绿色棒棒糖”，老师说：“嗯，你做的是白色棒棒糖”。她说：“我做的是白色棉花糖”。随后，她将揉好的白色黏土开始往木棍的顶端粘贴。单面粘贴牢固后，她将黏土向后拨，使其整个包围在木棍头的位置上，做好后她又将作品并排摆上刚才的作品旁边，她指着三个木棍高兴地对老师说：“这是三个了”。随后，她拿起超轻黏土盒准备继续进行“棒棒糖”制作游戏......\n"
    )

    completion = client.chat.completions.create(
        model="qwen-max-latest",
        messages=[{'role': 'user', 'content': prompt}],
    )
    return completion.model_dump_json()

def cleanup_old_files():
    """清理超过24小时的关键帧文件"""
    import time
    current_time = time.time()
    for session_dir in os.listdir(app.config['KEYFRAMES_FOLDER']):
        dir_path = os.path.join(app.config['KEYFRAMES_FOLDER'], session_dir)
        if os.path.isdir(dir_path):
            dir_time = os.path.getctime(dir_path)
            if current_time - dir_time > 600:  # 24小时 = 86400秒
                shutil.rmtree(dir_path, ignore_errors=True)

if __name__ == "__main__":
    cleanup_old_files()
    app.run(host="0.0.0.0", port=8080)
